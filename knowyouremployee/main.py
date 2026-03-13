from flask import Flask, request, redirect, url_for, render_template_string, flash
import pyodbc
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = "kye-secret-key"

# =========================
# CONFIG
# =========================
SERVER = "localhost"
DB_NAME = "KnowYourEmployeeDB"

EXCEL_FILE_PATH = r"C:\Users\LENOVO\OneDrive - JAY SWITCHES INDIA PVT LTD\know your Employee\Sheet- Know your Employee.xlsx"

def connect(database: str) -> pyodbc.Connection:
    conn_str = (
        "DRIVER={ODBC Driver 18 for SQL Server};"
        f"SERVER={SERVER};"
        f"DATABASE={database};"
        "Trusted_Connection=yes;"
        "Encrypt=yes;"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str)


# =========================
# HTML TEMPLATE
# =========================
FORM_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Know Your Employee Form</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        body {
            font-family: Arial, sans-serif;
            background: #f4f6f9;
            margin: 0;
            padding: 20px;
        }
        .container {
            max-width: 1100px;
            margin: auto;
            background: #fff;
            padding: 24px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.08);
        }
        h1, h2 {
            color: #1f3b5b;
        }
        .section {
            margin-bottom: 28px;
            padding: 18px;
            border: 1px solid #ddd;
            border-radius: 8px;
            background: #fafafa;
        }
        .grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
            gap: 14px;
        }
        label {
            display: block;
            font-weight: bold;
            margin-bottom: 6px;
            font-size: 14px;
        }
        input, select, textarea {
            width: 100%;
            padding: 9px;
            border: 1px solid #bbb;
            border-radius: 6px;
            font-size: 14px;
            box-sizing: border-box;
        }
        textarea {
            min-height: 80px;
            resize: vertical;
        }
        .repeat-block {
            border: 1px dashed #999;
            padding: 14px;
            margin-bottom: 12px;
            border-radius: 8px;
            background: white;
        }
        .btn {
            padding: 10px 16px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            margin-right: 8px;
        }
        .btn-primary {
            background: #0078d4;
            color: white;
        }
        .btn-secondary {
            background: #e5e5e5;
            color: black;
        }
        .btn-danger {
            background: #c62828;
            color: white;
        }
        .flash {
            padding: 12px;
            margin-bottom: 15px;
            border-radius: 6px;
            font-weight: bold;
        }
        .flash.success {
            background: #d4edda;
            color: #155724;
        }
        .flash.error {
            background: #f8d7da;
            color: #721c24;
        }
        .small-note {
            font-size: 12px;
            color: #555;
            margin-top: 5px;
        }
    </style>
</head>
<body>
<div class="container">
    <h1>Know Your Employee Form</h1>

    {% with messages = get_flashed_messages(with_categories=true) %}
      {% if messages %}
        {% for category, message in messages %}
          <div class="flash {{ category }}">{{ message }}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}

    <form method="POST" action="{{ url_for('submit_form') }}">

        <div class="section">
            <h2>1. Basic Details</h2>
            <div class="grid">
                <div>
                    <label>Form Date</label>
                    <input type="date" name="form_date">
                </div>
                <div>
                    <label>Employee Code</label>
                    <input type="text" name="employee_code">
                </div>
                <div>
                    <label>Card No</label>
                    <input type="text" name="card_no">
                </div>
                <div>
                    <label>Employee Name *</label>
                    <input type="text" name="employee_name" required>
                </div>
                <div>
                    <label>Gender</label>
                    <select name="gender">
                        <option value="">Select</option>
                        <option>Male</option>
                        <option>Female</option>
                        <option>Other</option>
                    </select>
                </div>
                <div>
                    <label>Department</label>
                    <input type="text" name="department">
                </div>
                <div>
                    <label>Designation</label>
                    <input type="text" name="designation">
                </div>
                <div>
                    <label>Date of Joining</label>
                    <input type="date" name="date_of_joining">
                </div>
                <div>
                    <label>Date of Birth</label>
                    <input type="date" name="date_of_birth">
                </div>
                <div>
                    <label>Mobile No</label>
                    <input type="text" name="mobile_no">
                </div>
                <div>
                    <label>Alternate Mobile No</label>
                    <input type="text" name="alternate_mobile_no">
                </div>
                <div>
                    <label>Email ID</label>
                    <input type="email" name="email_id">
                </div>
                <div>
                    <label>Aadhaar No</label>
                    <input type="text" name="aadhaar_no">
                </div>
                <div>
                    <label>PAN No</label>
                    <input type="text" name="pan_no">
                </div>
                <div>
                    <label>ESIC No</label>
                    <input type="text" name="esic_no">
                </div>
                <div>
                    <label>EPF No</label>
                    <input type="text" name="epf_no">
                </div>
                <div>
                    <label>Blood Group</label>
                    <input type="text" name="blood_group">
                </div>
                <div>
                    <label>Religion/Caste</label>
                    <input type="text" name="religion_caste">
                </div>
                <div>
                    <label>Marital Status</label>
                    <select name="marital_status">
                        <option value="">Select</option>
                        <option>Married</option>
                        <option>Unmarried</option>
                    </select>
                </div>
                <div>
                    <label>Spouse Name</label>
                    <input type="text" name="spouse_name">
                </div>
                <div>
                    <label>Marriage Anniversary</label>
                    <input type="date" name="marriage_anniversary">
                </div>
                <div>
                    <label>Nominee Details</label>
                    <input type="text" name="nominee_details">
                </div>
                <div>
                    <label>Employee Signature Date</label>
                    <input type="date" name="employee_signature_date">
                </div>
            </div>
        </div>

        <div class="section">
            <h2>2. Address / Family / Skills</h2>
            <div class="grid">
                <div>
                    <label>Permanent Address</label>
                    <textarea name="permanent_address"></textarea>
                </div>
                <div>
                    <label>Correspondence Address</label>
                    <textarea name="correspondence_address"></textarea>
                </div>
                <div>
                    <label>Father's Name / Occupation</label>
                    <input type="text" name="father_name_occupation">
                </div>
                <div>
                    <label>Mother's Name / Occupation</label>
                    <input type="text" name="mother_name_occupation">
                </div>
                <div>
                    <label>Skills</label>
                    <textarea name="skills"></textarea>
                </div>
            </div>
        </div>

        <div class="section">
            <h2>3. Education Details</h2>
            <div id="education-container">
                <div class="repeat-block education-block">
                    <div class="grid">
                        <div>
                            <label>Qualification</label>
                            <input type="text" name="education_qualification[]">
                        </div>
                        <div>
                            <label>Board / University</label>
                            <input type="text" name="education_board[]">
                        </div>
                        <div>
                            <label>Year of Passing</label>
                            <input type="text" name="education_year[]">
                        </div>
                        <div>
                            <label>Remarks</label>
                            <input type="text" name="education_remarks[]">
                        </div>
                    </div>
                </div>
            </div>
            <button type="button" class="btn btn-secondary" onclick="addEducation()">+ Add Education</button>
        </div>

        <div class="section">
            <h2>4. Previous Experience</h2>
            <div id="experience-container">
                <div class="repeat-block experience-block">
                    <div class="grid">
                        <div>
                            <label>Organization Name</label>
                            <input type="text" name="experience_org[]">
                        </div>
                        <div>
                            <label>Designation</label>
                            <input type="text" name="experience_designation[]">
                        </div>
                        <div>
                            <label>Period From</label>
                            <input type="date" name="experience_from[]">
                        </div>
                        <div>
                            <label>Period To</label>
                            <input type="date" name="experience_to[]">
                        </div>
                        <div>
                            <label>Remarks</label>
                            <input type="text" name="experience_remarks[]">
                        </div>
                    </div>
                </div>
            </div>
            <button type="button" class="btn btn-secondary" onclick="addExperience()">+ Add Experience</button>
        </div>

        <div class="section">
            <h2>5. Children Details</h2>
            <div id="children-container">
                <div class="repeat-block child-block">
                    <div class="grid">
                        <div>
                            <label>Child Name</label>
                            <input type="text" name="child_name[]">
                        </div>
                        <div>
                            <label>Gender</label>
                            <select name="child_gender[]">
                                <option value="">Select</option>
                                <option>Male</option>
                                <option>Female</option>
                                <option>Other</option>
                            </select>
                        </div>
                        <div>
                            <label>Date of Birth</label>
                            <input type="date" name="child_dob[]">
                        </div>
                        <div>
                            <label>Age</label>
                            <input type="number" name="child_age[]">
                        </div>
                        <div>
                            <label>Education</label>
                            <input type="text" name="child_education[]">
                        </div>
                        <div>
                            <label>Remarks / School / College</label>
                            <input type="text" name="child_remarks[]">
                        </div>
                    </div>
                </div>
            </div>
            <button type="button" class="btn btn-secondary" onclick="addChild()">+ Add Child</button>
        </div>

        <div class="section">
            <h2>6. Bank Details</h2>
            <div class="grid">
                <div>
                    <label>Bank Name</label>
                    <input type="text" name="bank_name">
                </div>
                <div>
                    <label>Branch Name</label>
                    <input type="text" name="branch_name">
                </div>
                <div>
                    <label>Account No</label>
                    <input type="text" name="account_no">
                </div>
                <div>
                    <label>IFSC Code</label>
                    <input type="text" name="ifsc_code">
                </div>
                <div>
                    <label>Nominee Name</label>
                    <input type="text" name="nominee_name">
                </div>
                <div>
                    <label>Nominee Relation</label>
                    <input type="text" name="nominee_relation">
                </div>
            </div>
        </div>

        <button type="submit" class="btn btn-primary">Submit Form</button>
    </form>
</div>

<script>
function addEducation() {
    const container = document.getElementById("education-container");
    const block = document.querySelector(".education-block").cloneNode(true);
    block.querySelectorAll("input").forEach(input => input.value = "");
    container.appendChild(block);
}

function addExperience() {
    const container = document.getElementById("experience-container");
    const block = document.querySelector(".experience-block").cloneNode(true);
    block.querySelectorAll("input").forEach(input => input.value = "");
    container.appendChild(block);
}

function addChild() {
    const container = document.getElementById("children-container");
    const block = document.querySelector(".child-block").cloneNode(true);
    block.querySelectorAll("input").forEach(input => input.value = "");
    block.querySelectorAll("select").forEach(select => select.selectedIndex = 0);
    container.appendChild(block);
}
</script>
</body>
</html>
"""


# =========================
# HELPERS
# =========================
def to_none(value: str):
    if value is None:
        return None
    value = value.strip()
    return value if value else None


def ensure_excel_workbook():
    """
    If workbook does not exist, create it with required sheets and headers.
    If it exists, keep using it.
    """
    if os.path.exists(EXCEL_FILE_PATH):
        return

    os.makedirs(os.path.dirname(EXCEL_FILE_PATH), exist_ok=True)

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Employees sheet
    ws_emp = wb.create_sheet("Employees")
    ws_emp.append([
        "EmployeeID", "FormDate", "EmployeeCode", "CardNo", "EmployeeName", "Gender",
        "Department", "Designation", "DateOfJoining", "DateOfBirth", "MobileNo",
        "AlternateMobileNo", "EmailID", "AadhaarNo", "PANNo", "ESICNo", "EPFNo",
        "BloodGroup", "ReligionCaste", "FatherNameOccupation", "MotherNameOccupation",
        "MaritalStatus", "SpouseName", "MarriageAnniversary", "Skills",
        "PermanentAddress", "CorrespondenceAddress", "NomineeDetails",
        "EmployeeSignatureDate", "CreatedAt"
    ])

    # Education sheet
    ws_edu = wb.create_sheet("Education")
    ws_edu.append([
        "EmployeeID", "Qualification", "BoardUniversityName", "YearOfPassing", "Remarks", "CreatedAt"
    ])

    # Experience sheet
    ws_exp = wb.create_sheet("Experience")
    ws_exp.append([
        "EmployeeID", "OrganizationName", "Designation", "PeriodFrom", "PeriodTo", "Remarks", "CreatedAt"
    ])

    # Children sheet
    ws_child = wb.create_sheet("Children")
    ws_child.append([
        "EmployeeID", "ChildName", "Gender", "DateOfBirth", "Age", "Education",
        "RemarksSchoolCollege", "CreatedAt"
    ])

    # BankDetails sheet
    ws_bank = wb.create_sheet("BankDetails")
    ws_bank.append([
        "EmployeeID", "BankName", "BranchName", "AccountNo", "IFSCCode",
        "NomineeName", "NomineeRelation", "CreatedAt"
    ])

    wb.save(EXCEL_FILE_PATH)


def save_to_excel(employee_id: int, employee_data: dict, educations: list, experiences: list, children: list, bank_data: dict):
    ensure_excel_workbook()

    wb = load_workbook(EXCEL_FILE_PATH)

    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Employees
    ws_emp = wb["Employees"]
    ws_emp.append([
        employee_id,
        employee_data.get("form_date"),
        employee_data.get("employee_code"),
        employee_data.get("card_no"),
        employee_data.get("employee_name"),
        employee_data.get("gender"),
        employee_data.get("department"),
        employee_data.get("designation"),
        employee_data.get("date_of_joining"),
        employee_data.get("date_of_birth"),
        employee_data.get("mobile_no"),
        employee_data.get("alternate_mobile_no"),
        employee_data.get("email_id"),
        employee_data.get("aadhaar_no"),
        employee_data.get("pan_no"),
        employee_data.get("esic_no"),
        employee_data.get("epf_no"),
        employee_data.get("blood_group"),
        employee_data.get("religion_caste"),
        employee_data.get("father_name_occupation"),
        employee_data.get("mother_name_occupation"),
        employee_data.get("marital_status"),
        employee_data.get("spouse_name"),
        employee_data.get("marriage_anniversary"),
        employee_data.get("skills"),
        employee_data.get("permanent_address"),
        employee_data.get("correspondence_address"),
        employee_data.get("nominee_details"),
        employee_data.get("employee_signature_date"),
        created_at
    ])

    # Education
    ws_edu = wb["Education"]
    for edu in educations:
        ws_edu.append([
            employee_id,
            edu.get("qualification"),
            edu.get("board"),
            edu.get("year"),
            edu.get("remarks"),
            created_at
        ])

    # Experience
    ws_exp = wb["Experience"]
    for exp in experiences:
        ws_exp.append([
            employee_id,
            exp.get("org"),
            exp.get("designation"),
            exp.get("from"),
            exp.get("to"),
            exp.get("remarks"),
            created_at
        ])

    # Children
    ws_child = wb["Children"]
    for child in children:
        ws_child.append([
            employee_id,
            child.get("name"),
            child.get("gender"),
            child.get("dob"),
            child.get("age"),
            child.get("education"),
            child.get("remarks"),
            created_at
        ])

    # Bank
    ws_bank = wb["BankDetails"]
    ws_bank.append([
        employee_id,
        bank_data.get("bank_name"),
        bank_data.get("branch_name"),
        bank_data.get("account_no"),
        bank_data.get("ifsc_code"),
        bank_data.get("nominee_name"),
        bank_data.get("nominee_relation"),
        created_at
    ])

    wb.save(EXCEL_FILE_PATH)


def save_to_sql(employee_data: dict, educations: list, experiences: list, children: list, bank_data: dict) -> int:
    conn = connect(DB_NAME)
    cursor = conn.cursor()

    try:
        # Insert main employee
        cursor.execute("""
            INSERT INTO dbo.Employees (
                FormDate, EmployeeCode, CardNo, EmployeeName, Gender, Department,
                Designation, DateOfJoining, DateOfBirth, MobileNo, AlternateMobileNo,
                EmailID, AadhaarNo, PANNo, ESICNo, EPFNo, BloodGroup, ReligionCaste,
                FatherNameOccupation, MotherNameOccupation, MaritalStatus, SpouseName,
                MarriageAnniversary, Skills, PermanentAddress, CorrespondenceAddress,
                NomineeDetails, EmployeeSignatureDate
            )
            OUTPUT INSERTED.EmployeeID
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            employee_data.get("form_date"),
            employee_data.get("employee_code"),
            employee_data.get("card_no"),
            employee_data.get("employee_name"),
            employee_data.get("gender"),
            employee_data.get("department"),
            employee_data.get("designation"),
            employee_data.get("date_of_joining"),
            employee_data.get("date_of_birth"),
            employee_data.get("mobile_no"),
            employee_data.get("alternate_mobile_no"),
            employee_data.get("email_id"),
            employee_data.get("aadhaar_no"),
            employee_data.get("pan_no"),
            employee_data.get("esic_no"),
            employee_data.get("epf_no"),
            employee_data.get("blood_group"),
            employee_data.get("religion_caste"),
            employee_data.get("father_name_occupation"),
            employee_data.get("mother_name_occupation"),
            employee_data.get("marital_status"),
            employee_data.get("spouse_name"),
            employee_data.get("marriage_anniversary"),
            employee_data.get("skills"),
            employee_data.get("permanent_address"),
            employee_data.get("correspondence_address"),
            employee_data.get("nominee_details"),
            employee_data.get("employee_signature_date")
        ))

        employee_id = cursor.fetchone()[0]

        # Education
        for edu in educations:
            cursor.execute("""
                INSERT INTO dbo.EmployeeEducation (
                    EmployeeID, Qualification, BoardUniversityName, YearOfPassing, Remarks
                )
                VALUES (?, ?, ?, ?, ?)
            """, (
                employee_id,
                edu.get("qualification"),
                edu.get("board"),
                edu.get("year"),
                edu.get("remarks")
            ))

        # Experience
        for exp in experiences:
            cursor.execute("""
                INSERT INTO dbo.EmployeeExperience (
                    EmployeeID, OrganizationName, Designation, PeriodFrom, PeriodTo, Remarks
                )
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                employee_id,
                exp.get("org"),
                exp.get("designation"),
                exp.get("from"),
                exp.get("to"),
                exp.get("remarks")
            ))

        # Children
        for child in children:
            cursor.execute("""
                INSERT INTO dbo.EmployeeChildren (
                    EmployeeID, ChildName, Gender, DateOfBirth, Age, Education, RemarksSchoolCollege
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                employee_id,
                child.get("name"),
                child.get("gender"),
                child.get("dob"),
                child.get("age"),
                child.get("education"),
                child.get("remarks")
            ))

        # Bank
        cursor.execute("""
            INSERT INTO dbo.EmployeeBankDetails (
                EmployeeID, BankName, BranchName, AccountNo, IFSCCode, NomineeName, NomineeRelation
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            employee_id,
            bank_data.get("bank_name"),
            bank_data.get("branch_name"),
            bank_data.get("account_no"),
            bank_data.get("ifsc_code"),
            bank_data.get("nominee_name"),
            bank_data.get("nominee_relation")
        ))

        conn.commit()
        return employee_id

    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()


# =========================
# ROUTES
# =========================
@app.route("/", methods=["GET"])
def index():
    return render_template_string(FORM_HTML)


@app.route("/submit", methods=["POST"])
def submit_form():
    try:
        employee_data = {
            "form_date": to_none(request.form.get("form_date")),
            "employee_code": to_none(request.form.get("employee_code")),
            "card_no": to_none(request.form.get("card_no")),
            "employee_name": to_none(request.form.get("employee_name")),
            "gender": to_none(request.form.get("gender")),
            "department": to_none(request.form.get("department")),
            "designation": to_none(request.form.get("designation")),
            "date_of_joining": to_none(request.form.get("date_of_joining")),
            "date_of_birth": to_none(request.form.get("date_of_birth")),
            "mobile_no": to_none(request.form.get("mobile_no")),
            "alternate_mobile_no": to_none(request.form.get("alternate_mobile_no")),
            "email_id": to_none(request.form.get("email_id")),
            "aadhaar_no": to_none(request.form.get("aadhaar_no")),
            "pan_no": to_none(request.form.get("pan_no")),
            "esic_no": to_none(request.form.get("esic_no")),
            "epf_no": to_none(request.form.get("epf_no")),
            "blood_group": to_none(request.form.get("blood_group")),
            "religion_caste": to_none(request.form.get("religion_caste")),
            "father_name_occupation": to_none(request.form.get("father_name_occupation")),
            "mother_name_occupation": to_none(request.form.get("mother_name_occupation")),
            "marital_status": to_none(request.form.get("marital_status")),
            "spouse_name": to_none(request.form.get("spouse_name")),
            "marriage_anniversary": to_none(request.form.get("marriage_anniversary")),
            "skills": to_none(request.form.get("skills")),
            "permanent_address": to_none(request.form.get("permanent_address")),
            "correspondence_address": to_none(request.form.get("correspondence_address")),
            "nominee_details": to_none(request.form.get("nominee_details")),
            "employee_signature_date": to_none(request.form.get("employee_signature_date")),
        }

        # Education arrays
        education_qualifications = request.form.getlist("education_qualification[]")
        education_boards = request.form.getlist("education_board[]")
        education_years = request.form.getlist("education_year[]")
        education_remarks = request.form.getlist("education_remarks[]")

        educations = []
        for q, b, y, r in zip(education_qualifications, education_boards, education_years, education_remarks):
            if any([to_none(q), to_none(b), to_none(y), to_none(r)]):
                educations.append({
                    "qualification": to_none(q),
                    "board": to_none(b),
                    "year": to_none(y),
                    "remarks": to_none(r),
                })

        # Experience arrays
        experience_orgs = request.form.getlist("experience_org[]")
        experience_designations = request.form.getlist("experience_designation[]")
        experience_froms = request.form.getlist("experience_from[]")
        experience_tos = request.form.getlist("experience_to[]")
        experience_remarks = request.form.getlist("experience_remarks[]")

        experiences = []
        for org, desig, frm, to, rem in zip(
            experience_orgs, experience_designations, experience_froms, experience_tos, experience_remarks
        ):
            if any([to_none(org), to_none(desig), to_none(frm), to_none(to), to_none(rem)]):
                experiences.append({
                    "org": to_none(org),
                    "designation": to_none(desig),
                    "from": to_none(frm),
                    "to": to_none(to),
                    "remarks": to_none(rem),
                })

        # Children arrays
        child_names = request.form.getlist("child_name[]")
        child_genders = request.form.getlist("child_gender[]")
        child_dobs = request.form.getlist("child_dob[]")
        child_ages = request.form.getlist("child_age[]")
        child_educations = request.form.getlist("child_education[]")
        child_remarks = request.form.getlist("child_remarks[]")

        children = []
        for name, gender, dob, age, education, remarks in zip(
            child_names, child_genders, child_dobs, child_ages, child_educations, child_remarks
        ):
            if any([to_none(name), to_none(gender), to_none(dob), to_none(age), to_none(education), to_none(remarks)]):
                children.append({
                    "name": to_none(name),
                    "gender": to_none(gender),
                    "dob": to_none(dob),
                    "age": int(age) if to_none(age) else None,
                    "education": to_none(education),
                    "remarks": to_none(remarks),
                })

        bank_data = {
            "bank_name": to_none(request.form.get("bank_name")),
            "branch_name": to_none(request.form.get("branch_name")),
            "account_no": to_none(request.form.get("account_no")),
            "ifsc_code": to_none(request.form.get("ifsc_code")),
            "nominee_name": to_none(request.form.get("nominee_name")),
            "nominee_relation": to_none(request.form.get("nominee_relation")),
        }

        if not employee_data["employee_name"]:
            flash("Employee Name is required.", "error")
            return redirect(url_for("index"))

        employee_id = save_to_sql(employee_data, educations, experiences, children, bank_data)
        save_to_excel(employee_id, employee_data, educations, experiences, children, bank_data)

        flash(f"Form submitted successfully. EmployeeID = {employee_id}", "success")
        return redirect(url_for("index"))

    except Exception as e:
        flash(f"Error while saving form: {str(e)}", "error")
        return redirect(url_for("index"))


# =========================
# MAIN
# =========================
if __name__ == "__main__":
    ensure_excel_workbook()
    app.run(debug=True)